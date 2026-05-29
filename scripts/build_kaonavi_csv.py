#!/usr/bin/env python3
"""カオナビ労務 給与明細CSV 統合ビルダー（月次運用）。

入力:
  - 給与明細送付先メール一覧
  - 従業員マスタ
  - 月次給与計算CSV（支給/控除/勤怠）正社員＋パート
  - 通勤手当/お弁当代/社会保険のIMPORT用CSV
  - パート時給一覧
  - 銀行口座PDF
  - 有給休暇管理xlsx（事業所別）
"""
import csv, re, codecs, subprocess
from pathlib import Path
import openpyxl

DL = Path("/Users/ryotanomura/Downloads")
EMAIL_CSV = DL / "給与明細送付先一覧 - シート6.csv"
MASTER = DL / "(株)ノムラ化成_従業員リスト_20250321～ - 従業員リスト.csv"
S_SHIKYU = DL / "202605_給与明細計算(正社員)  - 【支給】正社員・契約社員.csv"
S_KOJO = DL / "202605_給与明細計算(正社員)  - 【控除】正社員・契約社員.csv"
S_KINTAI = DL / "202605_給与明細計算(正社員)  - 【勤怠】正社員・契約社員.csv"
P_SHIKYU = DL / "202605_給与計算(パート・アルバイト)  - 【支給】パート社員・アルバイト社員.csv"
P_KOJO = DL / "202605_給与計算(パート・アルバイト)  - 【控除】パート社員・アルバイト社員.csv"
P_KINTAI = DL / "202605_給与計算(パート・アルバイト)  - 【勤怠】パート社員・アルバイト社員.csv"
PART_HOURLY = DL / "給与管理一覧_202509～ - 給与一覧(パート) (1).csv"
YUKYU_FILES = [
    (DL / "Tochigi_有給休暇管理.xlsx", "Tochigi_集計"),
    (DL / "Iwatsuki・Osaka_有給休暇管理.xlsx", "Iwatsuki_集計"),
    (DL / "Iwatsuki・Osaka_有給休暇管理.xlsx", "Osaka_集計"),
    (DL / "Gunma_有給休暇管理 (1).xlsx", "Gunma_集計"),
]
TEMPLATE = DL / "records-csv-sample-1756283139772 のコピー - records-csv-sample-1756283139772.csv"
BANK_PDF = DL / "SSK101-01_20250822162247.pdf"
OUT = DL / "カオナビ労務_給与明細_202605.csv"

EXCLUDE_NAMES = {("野村","富士子"),("大嶋","祐二"),("山岸","孝至"),
                 ("阿久澤","通子"),("川上","千香子"),("渡邉","勇斗"),
                 ("千森","直之"),("鈴木","英信")}
EXTRA_ADD = [("小山","英尚","I045"), ("原口","愛実","I047")]

def num(s):
    if s is None: return ""
    s = str(s).strip().strip('"').replace(",", "").replace(" ", "").replace("　","")
    if s == "" or s == "-": return ""
    neg = s.startswith("▲") or s.startswith("-")
    s = s.lstrip("▲-")
    try:
        v = float(s)
        if neg: v = -v
        return str(int(v)) if v == int(v) else str(v)
    except ValueError:
        return s

def add(*vals):
    total = 0.0; any_v = False
    for v in vals:
        n = num(v)
        if n:
            try: total += float(n); any_v = True
            except: pass
    if not any_v: return ""
    return str(int(total)) if total == int(total) else str(total)

def index_csv(path, key="社員ID"):
    d = {}
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            k = (row.get(key) or "").strip()
            if k: d[k] = row
    return d

master = index_csv(MASTER)
s_shikyu = index_csv(S_SHIKYU); s_kojo = index_csv(S_KOJO); s_kintai = index_csv(S_KINTAI)
p_shikyu = index_csv(P_SHIKYU); p_kojo = index_csv(P_KOJO); p_kintai = index_csv(P_KINTAI)

# 時給テーブル
hourly = {}
with open(PART_HOURLY, encoding="utf-8-sig") as f:
    rows = list(csv.reader(f))
header_h = rows[1]
for r in rows[2:]:
    if not r or not r[0].strip(): continue
    hourly[r[0].strip()] = dict(zip(header_h, r))

# 有休残（集計xlsxシートから）
yukyu = {}  # sid -> (残日数, 残時間)
def parse_zantime(v):
    """残時間: 数値(時間) or 'HH:MM:SS' or datetime.time → 時間(float)文字列"""
    if v is None: return ""
    if hasattr(v, "hour"):  # datetime.time
        h = v.hour + v.minute/60 + v.second/3600
        return str(int(h)) if h == int(h) else f"{h:.2f}".rstrip("0").rstrip(".")
    s = str(v).strip()
    if not s: return ""
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0]) + int(parts[1])/60
            return str(int(h)) if h == int(h) else f"{h:.2f}".rstrip("0").rstrip(".")
        except: return s
    return num(s)

for path, sheet in YUKYU_FILES:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames: continue
    ws = wb[sheet]
    # ヘッダー行: 3行目 (社員ID,氏名,..,残日数,残時間)
    for row in ws.iter_rows(min_row=4, values_only=True):
        sid = row[0]
        if not sid: continue
        sid = str(sid).strip()
        if not sid: continue
        zan_day = row[13] if len(row) > 13 else None
        zan_hour = row[14] if len(row) > 14 else None
        yukyu[sid] = (num(zan_day), parse_zantime(zan_hour))

# 銀行PDF パース
bank_text = subprocess.check_output(
    ["python3","-c",
     "from pypdf import PdfReader\nr=PdfReader(r'''"+str(BANK_PDF)+"''')\nprint('\\n'.join(p.extract_text() for p in r.pages))"]
).decode()

HW2FW = str.maketrans({
    "ｱ":"ア","ｲ":"イ","ｳ":"ウ","ｴ":"エ","ｵ":"オ",
    "ｶ":"カ","ｷ":"キ","ｸ":"ク","ｹ":"ケ","ｺ":"コ",
    "ｻ":"サ","ｼ":"シ","ｽ":"ス","ｾ":"セ","ｿ":"ソ",
    "ﾀ":"タ","ﾁ":"チ","ﾂ":"ツ","ﾃ":"テ","ﾄ":"ト",
    "ﾅ":"ナ","ﾆ":"ニ","ﾇ":"ヌ","ﾈ":"ネ","ﾉ":"ノ",
    "ﾊ":"ハ","ﾋ":"ヒ","ﾌ":"フ","ﾍ":"ヘ","ﾎ":"ホ",
    "ﾏ":"マ","ﾐ":"ミ","ﾑ":"ム","ﾒ":"メ","ﾓ":"モ",
    "ﾔ":"ヤ","ﾕ":"ユ","ﾖ":"ヨ",
    "ﾗ":"ラ","ﾘ":"リ","ﾙ":"ル","ﾚ":"レ","ﾛ":"ロ",
    "ﾜ":"ワ","ｦ":"ヲ","ﾝ":"ン","ｯ":"ッ",
    "ｬ":"ャ","ｭ":"ュ","ｮ":"ョ",
    "ｰ":"ー","ﾞ":"゛","ﾟ":"゜",
})
def hw_compose(s):
    out = []
    for ch in s.translate(HW2FW):
        if ch == "゛" and out:
            base = out[-1]
            d = {"カ":"ガ","キ":"ギ","ク":"グ","ケ":"ゲ","コ":"ゴ","サ":"ザ","シ":"ジ","ス":"ズ","セ":"ゼ","ソ":"ゾ",
                 "タ":"ダ","チ":"ヂ","ツ":"ヅ","テ":"デ","ト":"ド","ハ":"バ","ヒ":"ビ","フ":"ブ","ヘ":"ベ","ホ":"ボ","ウ":"ヴ"}
            if base in d: out[-1] = d[base]; continue
        if ch == "゜" and out:
            base = out[-1]
            h = {"ハ":"パ","ヒ":"ピ","フ":"プ","ヘ":"ペ","ホ":"ポ"}
            if base in h: out[-1] = h[base]; continue
        out.append(ch)
    return "".join(out)

bank_by_furi = {}
lines_pdf = [l.rstrip() for l in bank_text.split("\n")]
for i, line in enumerate(lines_pdf):
    m = re.search(r"普通\s*[－-]+\s*(\d+)", line)
    if not m or i < 4: continue
    kouza = m.group(1)
    bm = re.match(r"(.+?)（\d+）", lines_pdf[i-2].strip())
    bsm = re.match(r"(.+?)（\d+）", lines_pdf[i-1].strip())
    furi_full = hw_compose(lines_pdf[i-3].strip())
    bank_by_furi[furi_full] = {
        "銀行名": bm.group(1) if bm else lines_pdf[i-2].strip(),
        "支店名": bsm.group(1) if bsm else lines_pdf[i-1].strip(),
        "預金種別": "普通", "口座番号": kouza, "名義（カタカナ）": furi_full,
    }

def normalize_kana(s):
    SMALL2BIG = str.maketrans({"ャ":"ヤ","ュ":"ユ","ョ":"ヨ","ッ":"ツ","ァ":"ア","ィ":"イ","ゥ":"ウ","ェ":"エ","ォ":"オ"})
    s = s.translate(SMALL2BIG).replace("-","ー").replace(" ","").replace("　","")
    DAKU = {"ガ":"カ","ギ":"キ","グ":"ク","ゲ":"ケ","ゴ":"コ","ザ":"サ","ジ":"シ","ズ":"ス","ゼ":"セ","ゾ":"ソ",
            "ダ":"タ","ヂ":"チ","ヅ":"ツ","デ":"テ","ド":"ト","バ":"ハ","ビ":"ヒ","ブ":"フ","ベ":"ヘ","ボ":"ホ",
            "パ":"ハ","ピ":"ヒ","プ":"フ","ペ":"ヘ","ポ":"ホ","ヴ":"ウ"}
    return "".join(DAKU.get(c,c) for c in s)

ENGLISH_BANK_MAP = {
    "NAKANO APARECIDA SATIKO": ("Sachiko","Nakano"),
    "NUNES VERONICA TOMIKO NAKASHIM": ("Tomiko","Nakashima"),
    "WAKIMOTO ELIANE MAIUMY": ("Mayumi","Wakimoto"),
}
bank_by_english = {}
for k, v in list(bank_by_furi.items()):
    ku = k.upper().strip()
    for eng, name in ENGLISH_BANK_MAP.items():
        if eng in ku or ku in eng:
            bank_by_english[name] = v
bank_by_norm = {normalize_kana(k): v for k, v in bank_by_furi.items()}

def find_bank(furi_master, sei="", mei=""):
    if (sei, mei) in bank_by_english: return bank_by_english[(sei, mei)]
    fm = hw_compose(furi_master)
    if fm in bank_by_furi: return bank_by_furi[fm]
    norm = normalize_kana(fm)
    if norm in bank_by_norm: return bank_by_norm[norm]
    for k, v in bank_by_norm.items():
        if k.startswith(norm) or norm.startswith(k) or (norm and norm in k):
            return v
    return None

def fmt_birth(s):
    s = (s or "").strip()
    if not s: return ""
    m = re.match(r"(\d{4})[/年-](\d{1,2})[/月-](\d{1,2})", s)
    return f"{int(m.group(1))}年{int(m.group(2))}月{int(m.group(3))}日" if m else s

def split_furi(s):
    s = (s or "").strip()
    parts = re.split(r"[ 　]+", s, maxsplit=1)
    return (parts[0], parts[1]) if len(parts)==2 else (s, "")

PREF_RE = re.compile(r"^(東京都|北海道|京都府|大阪府|.+?県)")
CITY_RE = re.compile(r"^(.+?[市区](?:.+?[町村])?|.+?郡.+?[町村])")
def split_addr(addr):
    addr = (addr or "").strip()
    if not addr: return "","","",""
    pm = PREF_RE.match(addr)
    if not pm: return "","","",addr
    pref = pm.group(1); rest = addr[len(pref):]
    cm = CITY_RE.match(rest)
    if not cm: return pref,"",rest,""
    city = cm.group(1); rest2 = rest[len(city):]
    m = re.match(r"^([^\s]*?[\d０-９]+(?:[-－―ー][\d０-９]+)*(?:番地?(?:[\d０-９]+号?)?|号)?)(.*)$", rest2)
    if m and m.group(1):
        return pref, city, m.group(1), m.group(2).strip()
    return pref, city, rest2, ""

name_to_sid = {}
for sid, row in master.items():
    name = (row.get("氏名") or "").strip()
    parts = re.split(r"[ 　]+", name, maxsplit=1)
    if len(parts) == 2:
        name_to_sid[(parts[0], parts[1])] = sid
        name_to_sid[(parts[1], parts[0])] = sid

with open(TEMPLATE, encoding="utf-8-sig") as f:
    header = next(csv.reader(f))

def build_row(sei, mei, email, sid_override=""):
    sid = sid_override or name_to_sid.get((sei, mei), "")
    m = master.get(sid, {})
    is_seishain = sid in s_shikyu
    is_part = sid in p_shikyu
    if is_seishain:
        shikyu, kojo, kintai = s_shikyu[sid], s_kojo.get(sid,{}), s_kintai.get(sid,{})
    elif is_part:
        shikyu, kojo, kintai = p_shikyu[sid], p_kojo.get(sid,{}), p_kintai.get(sid,{})
    else:
        shikyu, kojo, kintai = {}, {}, {}

    koyou_master = (m.get("雇用形態") or "").strip()
    koyou = koyou_master or ("正社員" if is_seishain else ("パート" if is_part else ""))
    keitai = "時給" if koyou in ("パート","アルバイト") else ("月給" if koyou else "")

    fsei, fmei = split_furi(m.get("ﾌﾘｶﾞﾅ", ""))
    mname = (m.get("氏名") or "").strip()
    mparts = re.split(r"[ 　]+", mname, maxsplit=1)
    if len(mparts) == 2 and (mparts[0], mparts[1]) != (sei, mei) and (mparts[1], mparts[0]) == (sei, mei):
        fsei, fmei = fmei, fsei

    pref, city, banchi, bldg = split_addr(m.get("住所", ""))
    bank = find_bank(m.get("ﾌﾘｶﾞﾅ", ""), sei, mei) or {}

    kosei = add(kojo.get("厚生年金保険"), kojo.get("子ども・子育て支援金"))
    kotsu = add(shikyu.get("非課税通勤手当"), shikyu.get("課税通勤手当"))
    zeikin = add(kojo.get("所得税"), kojo.get("住民税"))
    sonota_kojo = add(kojo.get("その他控除"), kojo.get("お弁当代"))
    chikoku_kaisu = add(kintai.get("遅刻日数"), kintai.get("早退日数"))
    chikoku_jikan = add(kintai.get("遅刻時間"), kintai.get("早退時間"))
    tokubetsu_nichi = add(kintai.get("特別休暇利用日数(有給)"), kintai.get("特別休暇利用日数(無給)"))

    jikyu = ""
    if is_part and sid in hourly:
        jikyu = num(hourly[sid].get("基本給",""))

    # 有給残
    zan_day, zan_hour = yukyu.get(sid, ("", ""))
    biko = ""
    if zan_hour and zan_hour not in ("0","0.0"):
        biko = f"時間有給残: {zan_hour}時間"

    out = {
        "姓": sei, "名": mei,
        "姓（ヨミガナ）": fsei, "名（ヨミガナ）": fmei,
        "メールアドレス": email,
        "電話番号": (m.get("電話番号") or "").strip(),
        "住所（郵便番号）": (m.get("郵便番号") or "").strip(),
        "住所（都道府県）": pref, "住所（市区町村）": city,
        "住所（丁目・番地）": banchi, "住所（建物名・部屋番号）": bldg,
        "生年月日": fmt_birth(m.get("生年月日")),
        "社員コード": sid,
        "事業所": "", "雇用形態": koyou, "給与形態": keitai,
        "締め日・支払い日": "",
        "銀行名": bank.get("銀行名",""), "支店名": bank.get("支店名",""),
        "預金種別": bank.get("預金種別",""), "口座番号": bank.get("口座番号",""),
        "名義（カタカナ）": bank.get("名義（カタカナ）",""),
        "総支給額": num(shikyu.get("総支給額")),
        "控除合計": num(kojo.get("控除合計")),
        "差引支給額": num(shikyu.get("差引支給額")),
        "日給": "", "勤務日数": num(kintai.get("出勤日数")),
        "基本給": num(shikyu.get("基本給")) if not is_part else "",
        "時給": jikyu, "勤務時間": num(kintai.get("給与計算時間")),
        "ベース給": "", "役員報酬": "", "勤怠控除合計": "",
        "所定労働日数": num(kintai.get("要出勤日数")),
        "出勤日数": num(kintai.get("出勤日数")),
        "有休取得日数": num(kintai.get("有給休暇利用日数")),
        "有休残日数": zan_day,
        "欠勤日数": num(kintai.get("欠勤日数")),
        "欠勤控除": "",
        "遅刻早退回数": chikoku_kaisu,
        "遅刻早退時間": chikoku_jikan,
        "遅刻早退控除": "",
        "特別休暇日数": tokubetsu_nichi,
        "特別休暇時間": num(kintai.get("特別休暇時間換算")),
        "残業手当合計": num(shikyu.get("時間外手当計")),
        "法定時間内残業": "", "法定時間内残業手当": "",
        "法定時間外残業": num(kintai.get("時間外労働")),
        "法定時間外残業手当": num(shikyu.get("時間外手当")),
        "法定休日労働時間": "", "法定休日労働手当": num(shikyu.get("休日手当")),
        "深夜労働時間": num(kintai.get("深夜時間外労働")),
        "深夜労働手当": num(shikyu.get("深夜時間外手当")),
        "法定時間外割増": "", "法定時間外割増手当": "",
        "時間外労働60時間超過分": "", "時間外労働60時間超過分手当": "",
        "手当合計": num(shikyu.get("手当合計")),
        "役職手当": num(shikyu.get("役職手当")),
        "資格手当": num(shikyu.get("資格手当")),
        "交通費": kotsu,
        "その他合計（支給）": "", "立替精算": "",
        "社会保険料合計": num(kojo.get("社会保険料合計")),
        "健康保険料": num(kojo.get("健康保険")),
        "介護保険料": num(kojo.get("介護保険")),
        "厚生年金保険料": kosei,
        "雇用保険料": num(kojo.get("雇用保険")),
        "税金合計": zeikin,
        "所得税": num(kojo.get("所得税")),
        "住民税": num(kojo.get("住民税")),
        "その他合計（控除）": sonota_kojo,
        "備考": biko,
        "住宅手当": num(shikyu.get("住宅手当")),
        "扶養手当": num(shikyu.get("扶養手当")),
        "職務手当": num(shikyu.get("職務手当")),
        "皆勤手当": num(shikyu.get("皆勤手当")),
        "営業手当": num(shikyu.get("営業手当")),
        "技術手当": "",
        "在宅勤務手当": num(shikyu.get("在宅勤務手当")),
        "休日手当": num(shikyu.get("休日手当")),
        "休日時間外手当": num(shikyu.get("休日時間外手当")),
        "深夜時間外手当": num(shikyu.get("深夜時間外手当")),
        "お弁当代": num(kojo.get("お弁当代")),
    }
    if sid in ("I001","I002"):
        out["役員報酬"] = num(shikyu.get("基本給"))
        out["基本給"] = ""
    return out

rows = []
with open(EMAIL_CSV, encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        sei = r["姓"].strip(); mei = r["名"].strip(); email = r["メールアドレス"].strip()
        if (sei, mei) in EXCLUDE_NAMES: continue
        rows.append(build_row(sei, mei, email))

for sei, mei, sid in EXTRA_ADD:
    m = master.get(sid, {})
    email = (m.get("メールアドレス") or "").strip()
    rows.append(build_row(sei, mei, email, sid_override=sid))

with open(OUT, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=header)
    w.writeheader()
    for row in rows:
        w.writerow({k: row.get(k,"") for k in header})

no_bank = [r for r in rows if not r["銀行名"]]
no_yukyu = [r for r in rows if not r["有休残日数"]]
with_hour = [r for r in rows if r["備考"]]
print(f"出力: {OUT}")
print(f"行数: {len(rows)}")
print(f"\n有休残日数なし: {len(no_yukyu)}名 ({[r['社員コード'] for r in no_yukyu]})")
print(f"時間有給残あり (備考に記載): {len(with_hour)}名")
for r in with_hour:
    print(f"  - {r['姓']} {r['名']} ({r['社員コード']}): {r['備考']}")
print(f"\n銀行情報なし: {len(no_bank)}名 ({[r['社員コード'] for r in no_bank]})")
